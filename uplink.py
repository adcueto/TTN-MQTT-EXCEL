import time
import ttn
import openpyxl
import base64
from base64 import b64decode
global iRow, nGw;

app_id = "beacons"  #The app name
access_key = "ttn-account-v2.ErABTkHwaF7QCVuIY0ScHh9ezPnMCaUnwhcD8DqPIo8" # The

dest_filename = "LoRa.xlsx"  #The name of File Excel

doc = openpyxl.load_workbook(filename=dest_filename) #Open File Excel
gSheet = doc.worksheets[0]  #"general" Sheet
iRow=3 #begin of row third

#================= Function Callback ===========================================
def uplink_callback(msg, client):
    global iRow, nGw; #variable definition
    print("Received uplink from ", msg.dev_id)
    print("Battery",msg.payload_fields.batV)
    print("Frequency",msg.metadata.frequency)
    iColGw=21
    nGw = len(msg.metadata.gateways) # gateways number

    #===== General processing ==================================================
    gSheet.cell(row=iRow, column=1).value = msg.app_id
    gSheet.cell(row=iRow, column=2).value = msg.dev_id
    gSheet.cell(row=iRow, column=3).value = msg.hardware_serial
    gSheet.cell(row=iRow, column=4).value = msg.port
    gSheet.cell(row=iRow, column=5).value = msg.counter
    gSheet.cell(row=iRow, column=6).value = b64decode(msg.payload_raw).hex() #convert payload from base64 to hexadecimal
    #===== Payload processing ==================================================
    gSheet.cell(row=iRow, column=7).value = msg.payload_fields.batV
    gSheet.cell(row=iRow, column=8).value = msg.payload_fields.fixFailed
    gSheet.cell(row=iRow, column=9).value = msg.payload_fields.inTrip
    gSheet.cell(row=iRow, column=10).value = msg.payload_fields.headingDeg
    gSheet.cell(row=iRow, column=11).value = msg.payload_fields.latitudeDeg
    gSheet.cell(row=iRow, column=12).value = msg.payload_fields.longitudeDeg
    gSheet.cell(row=iRow, column=13).value = msg.payload_fields.speedKmph
    #====== Metadata processing ===============================================
    gSheet.cell(row=iRow, column=14).value = msg.metadata.time
    gSheet.cell(row=iRow, column=15).value = msg.metadata.frequency
    gSheet.cell(row=iRow, column=16).value = msg.metadata.modulation
    gSheet.cell(row=iRow, column=17).value = msg.metadata.data_rate
    gSheet.cell(row=iRow, column=18).value = msg.metadata.airtime
    gSheet.cell(row=iRow, column=19).value = msg.metadata.coding_rate
    gSheet.cell(row=iRow, column=20).value = nGw # gateways number
    #====== Gateways processing ================================================
    for iGw in range(nGw):
        gSheet.cell(row=iRow, column=iColGw+0).value = msg.metadata.gateways[iGw].gtw_id
        gSheet.cell(row=iRow, column=iColGw+1).value = msg.metadata.gateways[iGw].timestamp
        gSheet.cell(row=iRow, column=iColGw+2).value = msg.metadata.gateways[iGw].time
        gSheet.cell(row=iRow, column=iColGw+3).value = msg.metadata.gateways[iGw].channel
        gSheet.cell(row=iRow, column=iColGw+4).value = msg.metadata.gateways[iGw].rssi
        gSheet.cell(row=iRow, column=iColGw+5).value = msg.metadata.gateways[iGw].snr
        #gSheet.cell(row=iRow, column=iColGw+6).value = msg.metadata.gateways[iGw].latitude
        #gSheet.cell(row=iRow, column=iColGw+7).value = msg.metadata.gateways[iGw].longitude
        iColGw+=8; #column increment
    #====== End processing =====================================================
    doc.save(filename = dest_filename) #save file
    print("File updated")
    iRow+=1 #row increment

#================= end Function=================================================

handler = ttn.HandlerClient(app_id, access_key)
# using mqtt client
mqtt_client = handler.data()
mqtt_client.set_uplink_callback(uplink_callback)
mqtt_client.connect()
time.sleep(50000) # duration time
mqtt_client.close() #close mqtt client
